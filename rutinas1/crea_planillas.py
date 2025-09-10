from eval_insert import txt_a_df , inserta_archivo, ejecutasqlarch, ejecutasql, hace_conexion , cierra_conexion
from admin_df_prueba2 import convertir_a_json
from admin_4 import generar_html_docente, generar_html_estudiantes
from agrega_registros import agregar_registros
from datetime import datetime, timezone
import pandas as pd
import psycopg2
from psycopg2 import sql
from crea_pptx import creappt
import xlsxwriter
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname
import xlwings as xw
from openpyxl.workbook.protection import WorkbookProtection
import os


    

def copiasPlanillas(archivox, cod_asig, carpeta, sufijo, idseccion, fil , col):
    raiz, extension = os.path.splitext(archivox)
    path_base='C:/sudcraultra/Consultas/'

    archivo = open(path_base + 'listado_planillas.sql', "r")
    sql= archivo.read()
    if idseccion == "":
        sql = sql.replace("[seccion]", '')
    else:
        sql = sql.replace("[seccion]", 'AND s.id_seccion = ' + idseccion)

    sql = sql.replace("[cod_asig]", cod_asig)
    df=ejecutasql(sql)
    i=1
    total_registros = len(df)
    with tqdm(total=total_registros, desc=f"Progreso planillas {cod_asig}-{sufijo}") as pbar:
        for row in df.itertuples():
            archivo = open(path_base + 'listado_alumnos_planillas.sql', "r")
            sql= archivo.read()
            sql = sql.replace("[id_seccion]", str(row.id_seccion))
            dfa=ejecutasql(sql)
            ruta_completa = f"{carpeta}/{row.cod_programa}/{row.cod_sede}/{sufijo}/{cod_asig}"
  
            try:
                os.makedirs(ruta_completa)
            except Exception as e:
                m=0
            else:
                m=0
            finally:
                m=0
            
            seccion = f"{row.cod_sede}_{row.seccion}"
            new_planilla = f"{carpeta}/{row.cod_programa}/{row.cod_sede}/{sufijo}/{cod_asig}/{seccion}_{sufijo}{extension}"
            dfAplanilla(dfa,archivox,new_planilla, fil,col, seccion)
            pbar.update(1)



def dfAplanilla(df, excel_file ,new_excel_file,start_row, start_col, seccion):
    # Cargar el archivo Excel existente
    #wb = load_workbook(filename=excel_file, read_only=False)
    wb = load_workbook(filename=excel_file, read_only=False, keep_vba=True)
    if "SECCION" in wb.sheetnames:
        wb["SECCION"].title = seccion

    
    ws = wb.active
   
    # Obtener la posición de la celda (10,2)
    #start_row = 10
    #start_col = 2

    # Pega el DataFrame a partir de la celda (10,2)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
        for c_idx, value in enumerate(row, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)

            # Elimina las últimas filas

    for fila in range (start_row  + len(df), start_row + 60):  
         ws.row_dimensions[fila].hidden = True


    #while ws.max_row > 10 + len(df)-1:
        #ws.delete_rows(ws.max_row)
        #   ws.row_dimensions[ws.max_row].hidden = True
        


    # Proteger la hoja con la contraseña "arcdus"
    ws.protection.sheet = True
    ws.protection.set_password('arcdus')


    try:
        wsh = wb["Hoja1"]
        wsh.sheet_state = 'hidden'
    except KeyError:
        m=1
    
    
    
    
    wb.security = WorkbookProtection(workbookPassword = 'arcdus', lockStructure = True)
    

    # Guarda con un nombre diferente
    
    wb.save(new_excel_file)
    wb.close()

def dfAplanilla2(df, excel_file, new_excel_file, start_row, start_col):
    # Cargar el archivo Excel existente
    #wb = load_workbook(filename=excel_file, read_only=False)
    wb = load_workbook(filename=excel_file, read_only=False, keep_vba=True)
    ws = wb.active
   
    # Obtener la posición de la celda (10,2)
    #start_row = 10
    #start_col = 2

    # Pega el DataFrame a partir de la celda (10,2)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
        for c_idx, value in enumerate(row, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)

            # Elimina las últimas filas

    #for fila in range (start_row  + len(df), start_row + 60):  
    #     ws.row_dimensions[fila].hidden = True


    #while ws.max_row > 10 + len(df)-1:
        #ws.delete_rows(ws.max_row)
        #   ws.row_dimensions[ws.max_row].hidden = True
        


    # Proteger la hoja con la contraseña "arcdus"
    ws.protection.sheet = True
    ws.protection.set_password('arcdus')

    """
    try:
        wsh = wb["Hoja1"]
        wsh.sheet_state = 'hidden'
    except KeyError:
        print("La hoja 'Hoja1' no existe en el archivo Excel.")
    """
    
    
    
    wb.security = WorkbookProtection(workbookPassword = 'arcdus', lockStructure = True)
    

    # Guarda con un nombre diferente
    
    wb.save(new_excel_file)
    wb.close()


if __name__ == "__main__":

# prueba nuevas planillas
    if 1==2:
        cod_asig='MAT2120'
        prueba=2
        ruta='C:/sudcraultra_access/SISTEMA/planillas_base/'
        archivo=f'{cod_asig}-2024002-{prueba}.xlsm'
        rutaArchivo= ruta + archivo
        carp = 'C:/Users/lgutierrez/OneDrive - Fundacion Instituto Profesional Duoc UC/SUDCRA/PLANILLAS/TESTPLANILLAS'
        sufijo='P2'
        copiasPlanillas(rutaArchivo,cod_asig,carp,sufijo,"", 7,2)   
    
   
# MATEMÁTICA
    if 1==1:
        cod_asig='MAT3111'
        prueba=1
        ruta='C:/sudcraultra_access/SISTEMA/planillas_base/'
        archivo=f'{cod_asig}-2025002-{prueba}.xlsm'
        rutaArchivo= ruta + archivo
        carp = 'C:/Users/lgutierrez/Fundacion Instituto Profesional Duoc UC/Docentes Programa Matemática DUOC UC - PLANILLAS/MAT20252'
        sufijo='P1'
        copiasPlanillas(rutaArchivo,cod_asig,carp,sufijo,"", 7,2)   
    
 # EMPRENDIMIENTO   
    if 1==2 :
        cod_asig='EMP1701'
        prueba=4
        ruta='C:/sudcraultra_access/SISTEMA/planillas_base/'
        archivo=f'{cod_asig}-2024002-{prueba}.xlsm'
        rutaArchivo= ruta + archivo
        carp = 'C:/Users/lgutierrez/OneDrive - Fundacion Instituto Profesional Duoc UC/SUDCRA/PLANILLAS/EMP20242'
        sufijo='ET'
        copiasPlanillas(rutaArchivo,cod_asig,carp,sufijo,"", 7,2)
 # ETICA 
    if 1==2:
        cod_asig='PFC050'
        prueba=3
        ruta='C:/sudcraultra_access/SISTEMA/planillas_base/'
        archivo=f'{cod_asig}_{prueba}.xlsm'
        rutaArchivo= ruta + archivo
        carp = 'C:/Users/lgutierrez/OneDrive - Fundacion Instituto Profesional Duoc UC/SUDCRA/PLANILLAS/EYFC20242'
        sufijo='ET'
        copiasPlanillas(rutaArchivo,cod_asig,carp,sufijo,"", 10,2)


# LENGUAJE 
    if 1==0:
        cod_asig='PLC2110'
        prueba=1
        ruta='C:/sudcraultra_access/SISTEMA/planillas_base/'
        archivo=f'{cod_asig}-2025002-{prueba}.xlsm'
        rutaArchivo= ruta + archivo
        carp = 'C:/Users/lgutierrez/OneDrive - Fundacion Instituto Profesional Duoc UC/SUDCRA/PLANILLAS/LEN20252'
        sufijo='EV1'
        copiasPlanillas(rutaArchivo,cod_asig,carp,sufijo,"", 7,2)  

# INGLES 
    if 1==2:
        cod_asig='PRU7100'
        prueba=1
        ruta='C:/sudcraultra_access/SISTEMA/planillas_base/'
        archivo=f'{cod_asig}-2025001-{prueba}.xlsm'
        rutaArchivo= ruta + archivo
        carp = 'C:/Users/lgutierrez/OneDrive - Fundacion Instituto Profesional Duoc UC/SUDCRA/PLANILLAS/ING20251'
        sufijo='ET_ORAL'
        copiasPlanillas(rutaArchivo,cod_asig,carp,sufijo,"", 7,2)   
    
    


    if 1==2:
        path_base='C:/sudcraultra/Consultas/'
        # Consula SQL tabla informes_secciones_pendientes
        sql="select distinct e.cod_asig from eval e join asignaturas asig on asig.cod_asig = e.cod_asig where cod_programa = 'ing' and num_prueba = 1"
        df=ejecutasql(sql)
        # Consulta SQL tabla 

        i=1
        total_registros = len(df)
        with tqdm(total=total_registros, desc="Progreso informes secciones") as pbar:
            for row in df.itertuples():

                cod_asig=row.cod_asig
                prueba=1
                print(cod_asig)
                ruta='C:/sudcraultra_access/SISTEMA/planillas_base/'
                archivo=f'{cod_asig}-2024002-{prueba}.xlsm'
                rutaArchivo= ruta + archivo
                carp = 'C:/Users/lgutierrez/OneDrive - Fundacion Instituto Profesional Duoc UC/SUDCRA/PLANILLAS/ING20242'
                sufijo='ET_ORAL'
                copiasPlanillas(rutaArchivo,cod_asig,carp,sufijo,"", 7,2)  
                pbar.update(1)