from eval_insert import txt_a_df , inserta_archivo, ejecutasqlarch, ejecutasql, hace_conexion , cierra_conexion
from admin_df_prueba2 import convertir_a_json
from admin_4 import generar_html_docente, generar_html_estudiantes
from agrega_registros import agregar_registros
from datetime import datetime, timezone
import pandas as pd
import psycopg2
from psycopg2 import sql
from crea_pptx import creappt
import xlsxwriter
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook.protection import WorkbookProtection
import os

# Parámetros globales
ANIO = 2025
PERIODO = '002'
CARPETAONDRIVE = 'ING20252'  # MAT20252 - LEN20252 - IMR20252

def copiasPlanillas(archivox, cod_asig, carpeta, sufijo, idseccion, fil, col):
    path_base = 'C:/sudcraultra/Consultas/'
    raiz, extension = os.path.splitext(archivox)

    try:
        archivo = open(path_base + 'listado_planillas.sql', "r")
        sql = archivo.read()
        archivo.close()
    except Exception as e:
        print(f"❌ Error al leer listado_planillas.sql: {e}")
        return

    if idseccion == "":
        sql = sql.replace("[seccion]", '')
    else:
        sql = sql.replace("[seccion]", 'AND s.id_seccion = ' + idseccion)

    sql = sql.replace("[cod_asig]", cod_asig)
    try:
        df = ejecutasql(sql)
    except Exception as e:
        print(f"❌ Error al ejecutar SQL principal para {cod_asig}: {e}")
        return

    total_registros = len(df)
    if total_registros == 0:
        tqdm.write(f"⚠️ No hay registros para {cod_asig}-{sufijo}")
        return

    with tqdm(total=total_registros, desc=f"Progreso planillas {cod_asig}-{sufijo}") as pbar:
        for row in df.itertuples():
            try:
                archivo = open(path_base + 'listado_alumnos_planillas.sql', "r")
                sql = archivo.read()
                archivo.close()
                sql = sql.replace("[id_seccion]", str(row.id_seccion))
                dfa = ejecutasql(sql)
            except Exception as e:
                tqdm.write(f"❌ Error obteniendo alumnos para {cod_asig} sección {row.id_seccion}: {e}")
                pbar.update(1)
                continue

            ruta_completa = f"{carpeta}/{row.cod_programa}/{row.cod_sede}/{sufijo}/{cod_asig}"
            try:
                os.makedirs(ruta_completa, exist_ok=True)
            except Exception as e:
                tqdm.write(f"⚠️ No se pudo crear carpeta {ruta_completa}: {e}")

            seccion = f"{row.cod_sede}_{row.seccion}"
            new_planilla = f"{ruta_completa}/{seccion}_{sufijo}{extension}"

            try:
                dfAplanilla(dfa, archivox, new_planilla, fil, col, seccion)
            except Exception as e:
                tqdm.write(f"❌ Error creando planilla {new_planilla}: {e}")
            finally:
                pbar.update(1)

def dfAplanilla(df, excel_file, new_excel_file, start_row, start_col, seccion):
    try:
        wb = load_workbook(filename=excel_file, read_only=False, keep_vba=True)
    except FileNotFoundError:
        raise FileNotFoundError(f"Archivo base no encontrado: {excel_file}")
    except Exception as e:
        raise Exception(f"Error al abrir {excel_file}: {e}")

    try:
        if "SECCION" in wb.sheetnames:
            wb["SECCION"].title = seccion

        ws = wb.active
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
            for c_idx, value in enumerate(row, start=start_col):
                ws.cell(row=r_idx, column=c_idx, value=value)

        for fila in range(start_row + len(df), start_row + 60):
            ws.row_dimensions[fila].hidden = True

        ws.protection.sheet = True
        ws.protection.set_password('arcdus')

        try:
            wsh = wb["Hoja1"]
            wsh.sheet_state = 'hidden'
        except KeyError:
            pass

        wb.security = WorkbookProtection(workbookPassword='arcdus', lockStructure=True)
        wb.save(new_excel_file)
    finally:
        wb.close()

def ejecutar_planillas_ingles(lista_cod_asig, ruta_base, carpeta_destino):
    for cod_asig in lista_cod_asig:
        for prueba, sufijo in [(1, 'ORAL'), (2, 'ESCRITO')]:
            archivo = f"{cod_asig}-{ANIO}{PERIODO}-{prueba}.xlsm"
            rutaArchivo = os.path.join(ruta_base, archivo)
            tqdm.write(f"Procesando {cod_asig}, prueba {prueba} ({sufijo})...")
            try:
                copiasPlanillas(rutaArchivo, cod_asig, carpeta_destino, sufijo, "", 7, 2)
            except FileNotFoundError:
                tqdm.write(f"❌ Archivo base no encontrado para {cod_asig}-{sufijo}: {rutaArchivo}")
            except Exception as e:
                tqdm.write(f"❌ Error inesperado procesando {cod_asig}-{sufijo}: {e}")

if __name__ == "__main__":
    lista_ingles = [
        'INI1131', 'INI3101', 'INI3111', 'INI5111', 'INI5112', 'INI5121',
        'INI5131', 'INI7101', 'INI7102', 'INI7111', 'INI7121', 'INI9101',
        'INU1101', 'INU2101', 'INU3100', 'INU3101', 'INU3121', 'INU4101',
        'INU4121', 'INU5100', 'INU5101', 'INU6101', 'OPT1131', 'OPT3131',
        'PRI6100', 'PRU6100', 'PRU7100'
    ]

    ruta_base = r'C:/sudcraultra_access/SISTEMA/planillas_base/'
    carpeta_destino = r'C:/Users/lgutierrez/OneDrive - Fundacion Instituto Profesional Duoc UC/SUDCRA/PLANILLAS/' + CARPETAONDRIVE

    ejecutar_planillas_ingles(lista_ingles, ruta_base, carpeta_destino)
