import pandas as pd

from eval_insert import txt_a_df , inserta_archivo, ejecutasqlarch, ejecutasql, hace_conexion , cierra_conexion
import datetime
import time
import xlrd
import openpyxl

def xls_a_df(ruta, anoperiodo , id_archivoleido):
    path_base='C:/sudcraultra/Consultas/'
    df2 = pd.DataFrame(columns=['rut', 'id_itemresp', 'id_archivoleido', 'linea_leida', 'reproceso', 'imagen', 'instante_forms', 'num_prueba', 'forma','grupo', 'cod_interno', 'registro_leido'])
    
    df1 = pd.read_excel(ruta, header=None)
    # Lee cada línea del archivo
        
    linea_leida = 1
    n=0
   
    asig = df1.iloc[0, 127]
   
    prueba = int(df1.iloc[0, 128])
    validaPlanilla = df1.iloc[0, 1]
    id_seccionp=df1.iloc[0, 2]
    
    sql=f"SELECT e.cod_asig, asig.cod_interno, i.item_orden, i.item_tipo from secciones s join eval e on e.cod_asig = s.cod_asig join item i on i.id_eval = e.id_eval join asignaturas asig on asig.cod_asig = e.cod_asig where id_seccion = {id_seccionp} and i.forma = 1 and e.num_prueba = {prueba} order by i.item_orden"
    dfo=ejecutasql(sql)

    asig=dfo.at[0, 'cod_interno']
    # Imprimiendo los valores
    try:
        id_eval = asig + '-2024001-' + str(prueba)
    except:
        asig=str(int(asig))
        id_eval = asig  + '-2024001-' + str(prueba) 

    
    
    
    columnas_deseadas = list(range(2, 64)) + [125,126]



    df = pd.read_excel(ruta, skiprows=8, usecols=columnas_deseadas)
   
    # Filtra las columnas que no están completamente vacías
    try:
        df_sin_filas_vacias = df.dropna(subset=['Grupo','Nota'])
    except:
        df_sin_filas_vacias = df.dropna(subset=['FORMA','Nota']) 
    #print(df_sin_filas_vacias)
    df_limpio = df_sin_filas_vacias.dropna(axis=1, how='all')
    #print(df_limpio)
    if not df_limpio.empty:
        fg=df_limpio.columns[1]
        i=0
    
        for indice, fila in df_limpio.iterrows():
            #    Divide la línea en una lista de valores
            linea_leida=indice+10
            try:
                rep= fila['Reproceso']
            except:
                rep= fila['R']

            if rep != True:
                rep = False

            if rep == 1:
                rep = True

            if fg=='Grupo':
                forma=1
                grupo=int(fila['Grupo'])
                
            else:
                formatemp=fila['FORMA']
                if formatemp == 'a' or formatemp =='A':
                    forma=int(1)
                elif formatemp == 'b' or formatemp =='B':
                    forma=int(2)
                elif formatemp == 'c' or formatemp =='C':
                    forma=int(3)
                elif formatemp == 'd' or formatemp =='D':
                    forma=int(4)
                elif formatemp == 'e' or formatemp =='E':
                    forma=int(5)
                else:    
                    forma=int(fila['FORMA'])
                
                grupo=1

            if grupo == '':
                grupo = 0


            item = 1
            try:
                rut = int(fila['RUT'])
            except:
                rut = fila['RUT']
            inicio = 2  # Índice de la tercera columna
            fin = len(df_limpio.columns) - 2  # Índice de la antepenúltima columna
            
            for nombre_columna, valor in fila.items():
                if nombre_columna != 'RUT' and nombre_columna != 'FORMA' and nombre_columna != 'Grupo' and nombre_columna != 'Nota' and nombre_columna != 'Reproceso' and nombre_columna != 'R':
                    orden=item-2
                    try:
                        regi = int(valor)
                        imag=''
                        fecha_hora_forms = datetime.datetime.now()
                        if dfo.at[item-3, 'item_tipo']=='DE':
                            id_itemresp = anoperiodo + str(asig) + str(prueba).zfill(2) + str(forma).zfill(2) + str(orden).zfill(3)
                        else:
                            id_itemresp = anoperiodo + str(asig) + str(prueba).zfill(2) + str(forma).zfill(2) + str(orden).zfill(3) + str(int(regi))
                        nueva_fila = [rut, id_itemresp, id_archivoleido, linea_leida, rep, imag, fecha_hora_forms, prueba, forma, grupo , asig, valor]
                        df2.loc[n] = nueva_fila
                        nueva_fila = None
                    except Exception as e:
                        print('item: ' +  str(item))
                        print(e)
                n+= 1
                item += 1
            
        
        
   
    return df2

def xlsnew_a_df(ruta, id_archivoleido):
    path_base='C:/sudcraultra/Consultas/'
    df2 = pd.DataFrame(columns=['rut', 'id_itemresp', 'id_archivoleido', 'linea_leida', 'reproceso', 'imagen', 'instante_forms', 'num_prueba', 'forma','grupo', 'cod_interno', 'registro_leido'])
    
    df1 = pd.read_excel(ruta, sheet_name='datos')
    df1 = df1[['item_orden', 'tipo']]
    #print(df1)
    df_eval = pd.read_excel(ruta, sheet_name='Eval')
    df_eval = df_eval[['ano', 'periodo', 'num_prueba', 'cod_interno']]
    #print(df_eval )
    # Lee cada línea del archivo
        
    linea_leida = 1
    n=0
   
    asig = df_eval.loc[0, 'cod_interno']
    prueba = int(df_eval.loc[0, 'num_prueba'])
    ano = int(df_eval.loc[0, 'ano'])
    periodo = int(df_eval.loc[0, 'periodo'])
    anoperiodo = str(ano) + str(periodo).zfill(3)
  
    df = pd.read_excel(ruta,  usecols="C:AL", skiprows=5, nrows=60)
    df.columns.values[0] = 'RUT'
    df = df.dropna(subset=['Forma'])
    df_limpio = df.dropna(axis=1, how='all')
    
 
    #print(df_limpio)
    if not df_limpio.empty:
      
    
        for indice, fila in df_limpio.iterrows():
            #    Divide la línea en una lista de valores
            linea_leida=indice+6
            try:
                rep= fila['Reproceso']
            except:
                rep= fila['R']

            if rep == 'o':
                rep = False
            else:
                rep = True

            

            try:
                grupo=int(fila['Grupo'])
            except: 
                grupo=1
                   
            forma=int(fila['Forma'])
                
               

            if grupo == '':
                grupo = 0


            item = 1
            try:
                rut = int(fila['RUT'])
            except:
                rut = fila['RUT']
                      
            for nombre_columna, valor in fila.items():
                if nombre_columna != 'RUT' and nombre_columna != 'Forma' and nombre_columna != 'Grupo' and nombre_columna != 'Nota' and nombre_columna != 'Reproceso' and nombre_columna != 'Puntaje':
                    orden=item
                    try:
                        regi = int(valor)
                        imag=''
                        fecha_hora_forms = datetime.datetime.now()
                        if df1.at[item-1, 'tipo']=='DE':
                            id_itemresp = anoperiodo + str(asig) + str(prueba).zfill(2) + str(forma).zfill(2) + str(orden).zfill(3)
                        else:
                            id_itemresp = anoperiodo + str(asig) + str(prueba).zfill(2) + str(forma).zfill(2) + str(orden).zfill(3) + str(int(regi))
                        nueva_fila = [rut, id_itemresp, id_archivoleido, linea_leida, rep, imag, fecha_hora_forms, prueba, forma, grupo , asig, valor]
                        df2.loc[n] = nueva_fila
                        nueva_fila = None
                    except Exception as e:
                        print('item: ' +  str(item))
                        print(e)
                    item += 1
                n+= 1
                
            
        
        
   
    return df2


def xls_a_df2(ruta, anoperiodo , id_archivoleido):
    path_base='C:/sudcraultra/Consultas/'
    df2 = pd.DataFrame(columns=['rut', 'id_itemresp', 'id_archivoleido', 'linea_leida', 'reproceso', 'imagen', 'instante_forms', 'num_prueba', 'forma','grupo', 'cod_interno', 'registro_leido'])
    
    workbook =  openpyxl.load_workbook(ruta, read_only=True)
    sheet = workbook.active

    #df1 = pd.read_excel(ruta, header=None)
    # Lee cada línea del archivo
        
    linea_leida = 1
    n=0
   
    asig = sheet.cell(row=1, column=128).value
   
    prueba = int(sheet.cell(row=1, column=129).value)
    validaPlanilla = sheet.cell(row=1, column=2).value
    id_seccionp=sheet.cell(row=1, column=3).value
    workbook.close()
    time.sleep(1)
    workbook =  openpyxl.load_workbook("C:\\sudcraultra\\templates\\limpia.xlsx")
    time.sleep(1)
    workbook.close()
    sql=f"SELECT e.cod_asig, asig.cod_interno, i.item_orden, i.item_tipo from secciones s join eval e on e.cod_asig = s.cod_asig join item i on i.id_eval = e.id_eval join asignaturas asig on asig.cod_asig = e.cod_asig where id_seccion = {id_seccionp} and i.forma = 1 and e.num_prueba = {prueba} order by i.item_orden"
    dfo=ejecutasql(sql)
    # Imprimiendo los valores
    asig=dfo.at[0, 'cod_interno']
    try:
        id_eval = asig + '-2024001-' + str(prueba)
    except:
        asig=str(int(asig))
        id_eval = asig  + '-2024001-' + str(prueba) 

    

    
    columnas_deseadas = list(range(2, 64)) + [125,126]



    df = pd.read_excel(ruta, skiprows=8, usecols=columnas_deseadas)
   
    # Filtra las columnas que no están completamente vacías

    try:
        df_sin_filas_vacias = df.dropna(subset=['Grupo','Nota'])
    except:
        df_sin_filas_vacias = df.dropna(subset=['Forma','Nota']) 
    #print(df_sin_filas_vacias)
    df_limpio = df_sin_filas_vacias.dropna(axis=1, how='all')
    #print(df_limpio)
    if not df_limpio.empty:
        fg=df_limpio.columns[1]
        i=0
    
        for indice, fila in df_limpio.iterrows():
            #    Divide la línea en una lista de valores
            linea_leida=indice+10
            try:
                rep= fila['Reproceso']
            except:
                rep= fila['R']

            if rep != True:
                rep = False
                

            if fg=='Grupo':
                forma=1
                try:
                    grupo=int(fila['Grupo'])
                except:
                    grupo = 0
            else:
                forma=int(fila['FORMA'])
                
                grupo=1
            if grupo == '':
                grupo = 0

            item = 1
            rut = fila['RUT']
            inicio = 2  # Índice de la tercera columna
            fin = len(df_limpio.columns) - 2  # Índice de la antepenúltima columna
            
            for nombre_columna, valor in fila.items():
                if nombre_columna != 'RUT' and nombre_columna != 'FORMA' and nombre_columna != 'Grupo' and nombre_columna != 'Nota' and nombre_columna != 'Reproceso' and nombre_columna != 'R':
                    orden=item-2
                    try:
                        regi = int(valor)
                        imag=''
                        fecha_hora_forms = datetime.datetime.now()
                        if dfo.at[item-3, 'item_tipo']=='DE':
                            id_itemresp = anoperiodo + str(asig) + str(prueba).zfill(2) + str(forma).zfill(2) + str(orden).zfill(3)
                        else:
                            id_itemresp = anoperiodo + str(asig) + str(prueba).zfill(2) + str(forma).zfill(2) + str(orden).zfill(3) + str(int(regi))
                        nueva_fila = [rut, id_itemresp, id_archivoleido, linea_leida, rep, imag, fecha_hora_forms, prueba, forma, grupo , asig, valor]
                        df2.loc[n] = nueva_fila
                        nueva_fila = None
                    except Exception as e:
                        print('item: ' +  str(item))
                        print(e)
                n+= 1
                item += 1
            
        
        
   
    return df2


if __name__ == "__main__":
    archivo = 'C:/Users/lgutierrez/OneDrive - Fundacion Instituto Profesional Duoc UC/SUDCRA/PLANILLAS/TESTPLANILLAS/mat/VA/P2/MAT211/VI_MAT4140-002D_T4.xlsm'
    df= xlsnew_a_df(archivo, '1')
    print(df)